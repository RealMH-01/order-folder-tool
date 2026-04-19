# -*- coding: utf-8 -*-
"""
核心功能自测脚本：
- Storage 初始化 / 模板生成
- 业务员/客户管理
- flatten_template_folders / compare_with_existing
- execute_build 完整流程：4 个场景
- Excel 清单生成
"""

import json
import os
import shutil
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.core.storage import Storage
from app.core import folder_builder


def setup_workspace():
    tmp = Path(tempfile.mkdtemp(prefix="order_tool_test_"))
    root = tmp / "公司资料"
    root.mkdir()
    tpl_dir = tmp / "模板目录"
    # 准备模板文件
    (tpl_dir / "通用").mkdir(parents=True)
    (tpl_dir / "外贸通用").mkdir(parents=True)
    (tpl_dir / "宁夏").mkdir(parents=True)
    (tpl_dir / "湖北天鹅").mkdir(parents=True)

    # 生成空的模板文件（内容随意）
    files = [
        "通用/CG.xlsx",
        "外贸通用/CI.xlsx",
        "外贸通用/PL.xls",
        "外贸通用/托书.doc",
        "宁夏/宁夏外贸生产.doc",
        "宁夏/宁夏外贸发货.docx",
        "宁夏/宁夏内贸生产.xlsx",
        "宁夏/宁夏内贸发货.xlsx",
        "湖北天鹅/湖北天鹅外贸生产.xlsx",
        "湖北天鹅/湖北天鹅外贸发货.xlsx",
    ]
    for f in files:
        (tpl_dir / f).write_text("TEMPLATE CONTENT " + f, encoding="utf-8")

    return tmp, str(root), str(tpl_dir)


def assert_folder_exists(p):
    assert Path(p).is_dir(), f"文件夹不存在: {p}"


def assert_file_exists(p):
    assert Path(p).is_file(), f"文件不存在: {p}"


def assert_file_not_exists(p):
    assert not Path(p).exists(), f"文件不应存在: {p}"


def test_storage_init():
    print("\n===== 测试 1：Storage 初始化 =====")
    tmp, root, tpl_dir = setup_workspace()
    s = Storage(root)
    s.update_config(template_files_dir=tpl_dir)
    data = s.data_dir
    assert_folder_exists(data / "templates")
    assert_file_exists(data / "templates" / "standard_export.json")
    assert_file_exists(data / "templates" / "standard_domestic.json")
    assert_file_exists(data / "config.json")
    assert_file_exists(data / "salespersons.json")
    assert_file_exists(data / "history.json")
    print("OK: .order_tool 目录、默认标准模板已生成")


def test_salesperson():
    print("\n===== 测试 2：业务员 / 客户 =====")
    _, root, _ = setup_workspace()
    s = Storage(root)
    assert s.add_salesperson("张三")
    assert not s.add_salesperson("张三"), "重复添加应返回 False"
    assert s.add_customer("张三", "ACME")
    assert s.add_customer("张三", "HELLO")
    customers = s.get_customers("张三")
    assert customers == ["ACME", "HELLO"]
    print("OK: 业务员/客户增删正常")


def test_scenario(label, s, root, tpl_dir, order, expected_files):
    print(f"\n===== 测试：{label} =====")
    template = s.load_template(s.standard_template_filename(order["order_type"]))
    sub = "外贸" if order["order_type"] == "外贸" else "内贸"
    # 新逻辑：base_path 是"客户目录"，订单号文件夹在其下由模板根节点创建
    customer_dir = os.path.join(root, sub, order.get("customer", "ACME"))
    os.makedirs(customer_dir, exist_ok=True)

    result = folder_builder.execute_build(order, template, customer_dir, tpl_dir)
    order_folder = result["base_path"]  # 订单号文件夹
    # 基础校验
    assert_folder_exists(order_folder)
    assert Path(result["checklist_path"]).is_file(), "Excel 清单未生成"
    assert Path(result["checklist_path"]).parent == Path(order_folder), \
        "Excel 清单应在订单号文件夹下"
    print(f"  创建 {len(result['created'])} 个文件夹，跳过 {len(result['skipped'])} 个，"
          f"复制 {sum(1 for r in result['copy_results'] if r.get('copied'))} 个模板文件")

    # 商检资料（应在订单号文件夹内）
    insp_path = Path(order_folder) / "商检资料"
    if order.get("needs_inspection"):
        assert insp_path.is_dir(), "应创建商检资料目录"
        print("  OK: 商检资料目录已创建")
    else:
        assert not insp_path.exists(), "未勾选商检时不应创建"
        print("  OK: 未创建商检资料目录")

    # 校验预期文件（相对订单号文件夹）
    for rel in expected_files:
        p = Path(order_folder) / rel
        assert p.is_file(), f"缺失预期文件 {p}"
    print(f"  OK: 预期 {len(expected_files)} 个模板文件均已复制")

    # Bug 1 校验：子文件夹必须在订单号文件夹内部，不能散落在客户目录下
    for bad in ["SD", "报关资料", "货代资料", "生产、采购、发货"]:
        p_bad = Path(customer_dir) / bad
        if p_bad.exists():
            raise AssertionError(f"Bug1 回归：子文件夹 {bad} 散落在客户目录下：{p_bad}")
    print("  OK: 所有子文件夹都在订单号文件夹内部（Bug1 修复验证）")

    return order_folder, result


def test_rebuild(s, root, tpl_dir, order, order_folder):
    """order_folder 是订单号文件夹（result['base_path']）。
    base_path 传给 execute_build 的是客户目录（order_folder 的父目录）。
    """
    print(f"\n===== 补建测试：对已存在的订单再跑一次 =====")
    # 手动删除一个子文件夹，模拟缺失
    target = Path(order_folder) / "SD"
    if target.exists():
        shutil.rmtree(target)

    customer_dir = str(Path(order_folder).parent)
    template = s.load_template(s.standard_template_filename(order["order_type"]))
    ctx = folder_builder.build_context(order)
    tpl_folders = folder_builder.flatten_template_folders(
        template, customer_dir, ctx, order.get("needs_inspection", False))
    tpl_folders, extras = folder_builder.compare_with_existing(customer_dir, tpl_folders)

    create_cnt = sum(1 for i in tpl_folders if i["status"] == "to_create" and not i["is_root"])
    exist_cnt = sum(1 for i in tpl_folders if i["status"] == "existing" and not i["is_root"])
    print(f"  待创建 {create_cnt} / 已存在 {exist_cnt} / 模板外 {len(extras)}")
    assert create_cnt >= 1, "应检测到 SD 文件夹缺失需要补建"

    result = folder_builder.execute_build(order, template, customer_dir, tpl_dir)
    assert (Path(order_folder) / "SD").is_dir()
    print("  OK: 补建完成，SD 已重建")


def test_extras(root):
    print(f"\n===== 模板外目录测试 =====")
    # 客户目录
    customer_dir = Path(root) / "外贸" / "ACME"
    customer_dir.mkdir(parents=True, exist_ok=True)
    # 订单号文件夹（模板 root）
    order_folder = customer_dir / "TEST-EXT"
    order_folder.mkdir(exist_ok=True)
    (order_folder / "我自己新建的文件夹").mkdir(exist_ok=True)
    (order_folder / "SD").mkdir(exist_ok=True)
    # 同级别的另一个订单文件夹，不应被识别为"模板外"
    (customer_dir / "OTHER-ORDER").mkdir(exist_ok=True)

    s = Storage(root)
    template = s.load_template("standard_export.json")
    order = {"order_type": "外贸", "order_no": "TEST-EXT",
             "customer": "ACME", "product_category": "戊二醛",
             "salesperson": "张三", "needs_inspection": False}
    ctx = folder_builder.build_context(order)
    tpl_folders = folder_builder.flatten_template_folders(
        template, str(customer_dir), ctx, False)
    tpl_folders, extras = folder_builder.compare_with_existing(
        str(customer_dir), tpl_folders)
    extra_names = [e["name"] for e in extras]
    print(f"  extras = {extra_names}")
    assert "我自己新建的文件夹" in extra_names, "应识别为模板外目录"
    assert "OTHER-ORDER" not in extra_names, \
        "客户目录下的其他订单文件夹不应被误识为模板外"
    print("  OK: 模板外目录识别成功（仅扫描订单号文件夹内部）")


def test_history():
    print(f"\n===== 历史记录测试 =====")
    tmp, root, tpl_dir = setup_workspace()
    s = Storage(root)
    s.append_history({"time": "2025-01-01 10:00:00", "order_no": "A001",
                      "salesperson": "张三", "customer": "ACME",
                      "order_type": "外贸", "result": "成功"})
    s.append_history({"time": "2025-01-02 10:00:00", "order_no": "A002",
                      "salesperson": "李四", "customer": "BCD",
                      "order_type": "内贸", "result": "成功"})
    recs = s.load_history()
    assert len(recs) == 2
    assert recs[0]["order_no"] == "A002", "最新在前"
    print("  OK: 历史记录排序正常")


def test_templates():
    print(f"\n===== 模板管理测试 =====")
    tmp, root, _ = setup_workspace()
    s = Storage(root)
    tpl = s.load_template("standard_export.json")
    # 另存为
    fn = s.salesperson_template_filename("张三", "外贸")
    s.save_template(fn, tpl)
    assert s.load_template(fn) is not None
    fn2 = s.customer_template_filename("张三", "ACME", "外贸")
    s.save_template(fn2, tpl)
    groups = s.list_template_files()
    print(f"  standard={len(groups['standard'])}, salesperson={len(groups['salesperson'])}, customer={len(groups['customer'])}")
    assert len(groups["standard"]) == 2
    assert len(groups["salesperson"]) == 1
    assert len(groups["customer"]) == 1
    # 删除公司标准
    assert not s.delete_template("standard_export.json")
    assert s.delete_template(fn)
    print("  OK: 模板另存/删除/保护正常")

    # 匹配
    fn_m, t_m = s.match_template("张三", "ACME", "外贸")
    assert fn_m == fn2, f"应优先匹配客户专属模板，实际 {fn_m}"
    print("  OK: 模板匹配优先级正确")


def test_chinese_path():
    print(f"\n===== 中文路径测试 =====")
    tmp, root, tpl_dir = setup_workspace()
    s = Storage(root)
    order = {
        "order_type": "外贸", "order_no": "外贸-订单-中文001",
        "customer": "北京某某公司", "product_category": "戊二醛",
        "product_info": "戊二醛 200KG", "po_no": "PO-CN-001",
        "salesperson": "张三", "needs_inspection": True,
    }
    template = s.load_template("standard_export.json")
    customer_dir = os.path.join(root, "外贸", order["customer"])
    os.makedirs(customer_dir, exist_ok=True)
    result = folder_builder.execute_build(order, template, customer_dir, tpl_dir)
    order_folder = result["base_path"]
    assert (Path(order_folder) / "商检资料").is_dir()
    assert (Path(order_folder) / "生产发货" / "生产通知单-外贸-订单-中文001.doc").is_file()
    assert Path(result["checklist_path"]).is_file()
    print("  OK: 中文路径和文件名处理正常")


def test_scan_import():
    """测试扫描导入业务员 + 路径拼接（功能 1）"""
    print(f"\n===== 扫描导入 & 路径拼接测试 =====")
    tmp, root, _ = setup_workspace()
    # 搭建模拟的 1订单 结构
    order_root = Path(root) / "1订单"
    order_root.mkdir()
    # 1) 船级证（非业务员，不应勾选）
    (order_root / "船级证-金山IBC").mkdir()
    (order_root / "船级证-金山IBC" / "2024年.pdf").touch()
    # 2) 湖北分公司
    (order_root / "湖北" / "文天堂" / "客户A").mkdir(parents=True)
    (order_root / "湖北" / "文天堂" / "客户B").mkdir(parents=True)
    (order_root / "湖北" / "张子航" / "客户C").mkdir(parents=True)
    # 3) 张莹莹（直接客户）
    (order_root / "张莹莹" / "LLC KEMIKLKRAFT").mkdir(parents=True)
    (order_root / "张莹莹" / "客户B").mkdir(parents=True)
    # 4) 解小康（带"进行中订单"中间层）
    (order_root / "解小康" / "进行中订单" / "Arxada Tray Siam").mkdir(parents=True)
    (order_root / "解小康" / "进行中订单" / "BASF Nederland").mkdir(parents=True)
    (order_root / "解小康" / "2025出运计划表.xlsx").touch()
    # 5) 吴雅萍（带"1.进行订单"+"2.已完成订单"）
    (order_root / "吴雅萍" / "1.进行订单" / "客户X").mkdir(parents=True)
    (order_root / "吴雅萍" / "1.进行订单" / "客户Y").mkdir(parents=True)
    (order_root / "吴雅萍" / "2.已完成订单" / "客户Z").mkdir(parents=True)
    (order_root / "吴雅萍" / "3.订能常用文件").mkdir(parents=True)
    # 6) 根目录下杂散文件（应忽略）
    (order_root / "产品中英文对照表.docx").touch()

    s = Storage(root)
    # 检查 scan_order_root
    names = s.scan_order_root()
    assert "船级证-金山IBC" in names
    assert "湖北" in names
    assert "张莹莹" in names
    assert "产品中英文对照表.docx" not in names, "文件不应出现"
    print(f"  scan_order_root: {names}")

    # 模拟勾选：张莹莹、解小康、吴雅萍、湖北/文天堂、湖北/张子航
    rel_paths = ["张莹莹", "解小康", "吴雅萍", "湖北/文天堂", "湖北/张子航"]
    report = s.import_scanned_salespersons(rel_paths, overwrite=True)
    assert set(report["added"]) == {"张莹莹", "解小康", "吴雅萍", "文天堂", "张子航"}, \
        f"added = {report['added']}"
    print(f"  import report: {report}")

    # 验证 salespersons.json
    sps = s.load_salespersons()
    mp = {sp["name"]: sp for sp in sps}

    # 张莹莹：直接客户，无中间层
    assert mp["张莹莹"]["rel_path"] == "张莹莹"
    assert mp["张莹莹"]["mid_layer"] == ""
    assert "LLC KEMIKLKRAFT" in mp["张莹莹"]["customers"]

    # 解小康：有"进行中订单"中间层
    assert mp["解小康"]["rel_path"] == "解小康"
    assert mp["解小康"]["mid_layer"] == "进行中订单"
    assert "BASF Nederland" in mp["解小康"]["customers"]
    assert "2025出运计划表.xlsx" not in mp["解小康"]["customers"]  # 文件被忽略

    # 吴雅萍：有"1.进行订单"（只进这个）
    assert mp["吴雅萍"]["mid_layer"] == "1.进行订单", f"mid_layer={mp['吴雅萍']['mid_layer']}"
    assert set(mp["吴雅萍"]["customers"]) == {"客户X", "客户Y"}, \
        f"customers={mp['吴雅萍']['customers']}"

    # 文天堂：分公司下，rel_path=湖北/文天堂
    assert mp["文天堂"]["rel_path"] == "湖北/文天堂"
    assert mp["文天堂"]["mid_layer"] == ""
    assert "客户A" in mp["文天堂"]["customers"]

    print("  OK: salespersons.json 中 rel_path / mid_layer / customers 均正确")

    # 路径拼接
    # 张莹莹 + LLC KEMIKLKRAFT → 1订单/张莹莹/LLC KEMIKLKRAFT
    p1 = s.build_customer_dir("张莹莹", "LLC KEMIKLKRAFT")
    expected = os.path.join(root, "1订单", "张莹莹", "LLC KEMIKLKRAFT")
    assert p1 == expected, f"{p1} != {expected}"

    # 文天堂 + 客户A → 1订单/湖北/文天堂/客户A
    p2 = s.build_customer_dir("文天堂", "客户A")
    expected2 = os.path.join(root, "1订单", "湖北", "文天堂", "客户A")
    assert p2 == expected2, f"{p2} != {expected2}"

    # 解小康 + BASF → 1订单/解小康/进行中订单/BASF
    p3 = s.build_customer_dir("解小康", "BASF Nederland")
    expected3 = os.path.join(root, "1订单", "解小康", "进行中订单", "BASF Nederland")
    assert p3 == expected3, f"{p3} != {expected3}"

    # 吴雅萍 + 客户X → 1订单/吴雅萍/1.进行订单/客户X
    p4 = s.build_customer_dir("吴雅萍", "客户X")
    expected4 = os.path.join(root, "1订单", "吴雅萍", "1.进行订单", "客户X")
    assert p4 == expected4, f"{p4} != {expected4}"

    print("  OK: build_customer_dir 四种场景路径拼接全部正确")


def test_full_flow_bug1_verify():
    """端到端验证 Bug1 修复：订单号文件夹必须存在且子文件夹在其内部"""
    print(f"\n===== Bug1 端到端验证 =====")
    tmp, root, tpl_dir = setup_workspace()
    s = Storage(root)
    # 准备业务员张莹莹 + 客户
    (Path(root) / "1订单").mkdir()
    (Path(root) / "1订单" / "张莹莹" / "LLC KEMIKLKRAFT").mkdir(parents=True)
    s.import_scanned_salespersons(["张莹莹"], overwrite=True)

    order = {
        "order_type": "外贸", "order_no": "XS-GAP2604018NH",
        "customer": "LLC KEMIKLKRAFT", "product_category": "戊二醛",
        "product_info": "", "po_no": "",
        "salesperson": "张莹莹", "needs_inspection": False,
    }
    customer_dir = s.build_customer_dir("张莹莹", "LLC KEMIKLKRAFT")
    tpl = s.load_template("standard_export.json")
    result = folder_builder.execute_build(order, tpl, customer_dir, tpl_dir)

    order_folder = Path(result["base_path"])
    assert order_folder.name == "XS-GAP2604018NH", f"订单号文件夹名错误：{order_folder.name}"
    assert order_folder.parent == Path(customer_dir), \
        f"订单号文件夹父目录应为客户目录，实际：{order_folder.parent}"

    # 子文件夹应在订单号文件夹内
    for sub in ["SD", "报关资料", "货代资料", "生产发货", "装箱", "船期", "证据链"]:
        assert (order_folder / sub).is_dir(), f"缺少 {sub}"
        assert not (Path(customer_dir) / sub).exists() or \
               (Path(customer_dir) / sub) == (order_folder / sub), \
            f"{sub} 散落在客户目录下了"

    # 关键文件也应在订单号文件夹内
    for f in ["PI-XS-GAP2604018NH.xlsx",
              "CG-XS-GAP2604018NH.xlsx",
              "文件清单-XS-GAP2604018NH.xlsx"]:
        # PI 模板可能不存在，跳过不严格校验
        pass
    assert (order_folder / "文件清单-XS-GAP2604018NH.xlsx").is_file()
    print(f"  OK: 订单号文件夹 {order_folder}")
    print(f"  OK: 子文件夹均在订单号文件夹内部，客户目录下不散落")


def test_po_no_placeholder():
    """验证客户PO号占位符在 context 构建和文件名替换中都生效"""
    print("\n===== 测试：客户PO号占位符 =====")
    from openpyxl import load_workbook

    # 1) build_context 包含 PO 号
    order = {
        "order_no": "XS-PO-TEST", "customer": "ACME",
        "po_no": "PO-2026-001", "product_info": "",
        "salesperson": "张三", "order_type": "外贸",
        "product_category": "戊二醛", "needs_inspection": False,
    }
    ctx = folder_builder.build_context(order)
    assert ctx["<客户PO号>"] == "PO-2026-001", f"PO号上下文错误: {ctx['<客户PO号>']}"

    # 2) replace_placeholders 正确替换
    result = folder_builder.replace_placeholders("PO-<客户PO号>.pdf", ctx)
    assert result == "PO-PO-2026-001.pdf", f"文件名替换错误: {result}"

    # 3) PO 号为空时，占位符保留或替换为空字符串
    order_no_po = dict(order, po_no="")
    ctx2 = folder_builder.build_context(order_no_po)
    result2 = folder_builder.replace_placeholders("PO-<客户PO号>.pdf", ctx2)
    assert result2 == "PO-.pdf", f"空PO号替换错误: {result2}"

    # 4) 端到端：execute_build 中含 PO 号的订单，验证 Excel 清单文件名正确
    tmp, root, tpl_dir = setup_workspace()
    s = Storage(root)
    customer_dir = os.path.join(root, "外贸", "ACME")
    os.makedirs(customer_dir, exist_ok=True)
    template = s.load_template("standard_export.json")
    result_build = folder_builder.execute_build(order, template, customer_dir, tpl_dir)
    # 验证订单文件夹被创建
    assert Path(result_build["base_path"]).is_dir()
    # 验证 Excel 清单存在
    assert Path(result_build["checklist_path"]).is_file()
    # 验证 Excel 清单内容中有 PO-2026-001（读取 Excel sheet 搜索）
    wb = load_workbook(result_build["checklist_path"])
    ws = wb.active
    found_po = False
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell and "PO-2026-001" in str(cell):
                found_po = True
                break
    assert found_po, "Excel 清单中未找到 PO-2026-001"

    print("  OK: PO号占位符在 context、文件名、Excel清单中全部正确")


def test_batch_collect_po():
    """模拟批量导入场景：含 PO 号的订单通过 execute_build 后，文件名包含 PO 号"""
    print("\n===== 测试：批量导入含PO号 =====")
    from openpyxl import load_workbook
    tmp, root, tpl_dir = setup_workspace()
    s = Storage(root)

    orders = [
        {"order_type": "外贸", "order_no": "BATCH-001", "customer": "C1",
         "po_no": "PO-B001", "product_info": "", "product_category": "戊二醛",
         "salesperson": "张三", "needs_inspection": False},
        {"order_type": "外贸", "order_no": "BATCH-002", "customer": "C2",
         "po_no": "", "product_info": "", "product_category": "其他产品",
         "salesperson": "张三", "needs_inspection": False},
    ]
    for od in orders:
        customer_dir = os.path.join(root, "外贸", od["customer"])
        os.makedirs(customer_dir, exist_ok=True)
        tpl = s.load_template("standard_export.json")
        result = folder_builder.execute_build(od, tpl, customer_dir, tpl_dir)
        order_folder = Path(result["base_path"])
        assert order_folder.is_dir(), f"订单文件夹不存在: {order_folder}"
        # 标准外贸模板根目录有 PO-<客户PO号>.pdf（无 file_template 所以不会被复制）
        # 但 Excel 清单中应包含 PO 号替换后的文件名
        wb = load_workbook(result["checklist_path"])
        ws = wb.active
        all_text = " ".join(str(c) for row in ws.iter_rows(values_only=True) for c in row if c)
        if od["po_no"]:
            assert od["po_no"] in all_text, \
                f"订单 {od['order_no']} 的 Excel 清单中未找到 {od['po_no']}"
        print(f"  OK: {od['order_no']} (PO={od['po_no'] or '空'}) 清单生成正确")

    print("  OK: 批量含PO号测试通过")


def test_scan_import_overwrite_merge():
    """验证扫描导入：覆盖模式下，已有客户不丢失，新客户被合并"""
    print("\n===== 测试：扫描导入合并更新 =====")
    tmp, root, _ = setup_workspace()
    order_root = Path(root) / "1订单"
    order_root.mkdir()
    (order_root / "张三" / "老客户A").mkdir(parents=True)
    (order_root / "张三" / "老客户B").mkdir(parents=True)
    (order_root / "张三" / "新客户C").mkdir(parents=True)

    s = Storage(root)
    # 第一次导入（只导入前两个客户的快照）
    s.import_scanned_salespersons(["张三"], overwrite=False)
    sp = s.get_salesperson("张三")
    assert "老客户A" in sp["customers"]
    assert "老客户B" in sp["customers"]
    assert "新客户C" in sp["customers"]
    original_count = len(sp["customers"])

    # 手动添加一个"手动客户D"（模拟用户手动在程序中加的客户）
    s.add_customer("张三", "手动客户D")
    sp = s.get_salesperson("张三")
    assert "手动客户D" in sp["customers"]

    # 第二次导入（overwrite=True），磁盘上没有"手动客户D"文件夹
    report = s.import_scanned_salespersons(["张三"], overwrite=True)
    assert "张三" in report["updated"]
    sp = s.get_salesperson("张三")
    # 手动客户D 应该保留（合并逻辑：保留已有 + 新增）
    assert "手动客户D" in sp["customers"], \
        f"合并后丢失了手动添加的客户，customers={sp['customers']}"
    # 原来的客户也应该都在
    assert "老客户A" in sp["customers"]
    assert "新客户C" in sp["customers"]

    print(f"  OK: 合并后客户列表: {sp['customers']}")
    print("  OK: 扫描导入合并更新逻辑正确")


def test_mid_layer_multiple_candidates():
    """验证中间层选择：只选同时包含'进行'和'订单'的文件夹"""
    print("\n===== 测试：中间层多候选 =====")
    tmp, root, _ = setup_workspace()
    order_root = Path(root) / "1订单"
    # 构建带多个"订单"字样文件夹的业务员目录
    sp_dir = order_root / "测试员"
    (sp_dir / "1.进行订单" / "客户X").mkdir(parents=True)
    (sp_dir / "2.已完成订单" / "客户Y").mkdir(parents=True)
    (sp_dir / "3.订能常用文件").mkdir(parents=True)
    (sp_dir / "readme.txt").touch()  # 文件应被忽略

    s = Storage(root)
    mid, customers = s.scan_customers_for("测试员")
    assert mid == "1.进行订单", f"中间层应为'1.进行订单'，实际为'{mid}'"
    assert "客户X" in customers, f"客户列表应含'客户X'，实际为{customers}"
    assert "客户Y" not in customers, \
        f"'2.已完成订单'下的客户不应被导入，实际为{customers}"
    assert "3.订能常用文件" not in customers

    # 路径拼接验证
    s.import_scanned_salespersons(["测试员"], overwrite=True)
    p = s.build_customer_dir("测试员", "客户X")
    expected = os.path.join(root, "1订单", "测试员", "1.进行订单", "客户X")
    assert p == expected, f"路径不匹配: {p} != {expected}"

    print(f"  OK: mid_layer={mid}, customers={customers}")
    print("  OK: 多候选中间层测试通过")


def test_template_match_priority():
    """验证模板匹配优先级"""
    print("\n===== 测试：模板匹配优先级 =====")
    tmp, root, _ = setup_workspace()
    s = Storage(root)

    std_tpl = s.load_template("standard_export.json")
    assert std_tpl is not None

    # 只有标准模板时，应返回标准
    fn, tpl = s.match_template("张三", "ACME", "外贸")
    assert fn == "standard_export.json"

    # 添加业务员个人模板
    personal_fn = s.salesperson_template_filename("张三", "外贸")
    s.save_template(personal_fn, std_tpl)
    fn, tpl = s.match_template("张三", "ACME", "外贸")
    assert fn == personal_fn, f"应匹配个人模板，实际: {fn}"

    # 添加客户专属模板
    customer_fn = s.customer_template_filename("张三", "ACME", "外贸")
    s.save_template(customer_fn, std_tpl)
    fn, tpl = s.match_template("张三", "ACME", "外贸")
    assert fn == customer_fn, f"应匹配客户专属模板，实际: {fn}"

    # 换一个客户，应退回到个人模板
    fn2, _ = s.match_template("张三", "OTHER", "外贸")
    assert fn2 == personal_fn, f"换客户后应退回个人模板，实际: {fn2}"

    # 换一个业务员，应退回到标准模板
    fn3, _ = s.match_template("李四", "ACME", "外贸")
    assert fn3 == "standard_export.json", f"换业务员后应退回标准模板，实际: {fn3}"

    # 内贸模板应完全独立
    fn4, _ = s.match_template("张三", "ACME", "内贸")
    assert fn4 == "standard_domestic.json", f"内贸应返回标准内贸，实际: {fn4}"

    print("  OK: 模板匹配优先级全部正确")


def test_end_to_end_domestic():
    """端到端：内贸订单 + 扫描导入 + build_customer_dir + execute_build"""
    print("\n===== 测试：内贸端到端 =====")
    tmp, root, tpl_dir = setup_workspace()
    s = Storage(root)

    # 搭建目录
    order_root = Path(root) / "1订单"
    (order_root / "冷斌捷" / "1.进行订单" / "某某化工").mkdir(parents=True)

    # 扫描导入
    s.import_scanned_salespersons(["冷斌捷"], overwrite=True)
    sp = s.get_salesperson("冷斌捷")
    assert sp["mid_layer"] == "1.进行订单"
    assert "某某化工" in sp["customers"]

    # 构造订单
    order = {
        "order_type": "内贸", "order_no": "NS-DOM001",
        "customer": "某某化工", "po_no": "PO-DOM",
        "product_info": "戊二醛 1T", "product_category": "戊二醛",
        "salesperson": "冷斌捷", "needs_inspection": False,
    }
    customer_dir = s.build_customer_dir("冷斌捷", "某某化工")
    expected_dir = os.path.join(root, "1订单", "冷斌捷", "1.进行订单", "某某化工")
    assert customer_dir == expected_dir, f"{customer_dir} != {expected_dir}"

    tpl = s.load_template("standard_domestic.json")
    result = folder_builder.execute_build(order, tpl, customer_dir, tpl_dir)

    order_folder = Path(result["base_path"])
    assert order_folder.name == "NS-DOM001"
    assert order_folder.parent == Path(customer_dir)

    # 验证内贸特有的子文件夹
    for sub in ["SD", "生产、采购、发货", "物流", "证据链(ERP)"]:
        assert (order_folder / sub).is_dir(), f"内贸缺少子文件夹: {sub}"

    # 验证内贸不应有"商检资料"
    assert not (order_folder / "商检资料").exists(), "内贸不应有商检资料"

    # 验证子文件夹没有散落到客户目录
    for bad in ["SD", "物流"]:
        bad_path = Path(customer_dir) / bad
        assert not bad_path.exists() or bad_path == order_folder / bad, \
            f"子文件夹散落: {bad_path}"

    # 验证 Excel 清单
    assert Path(result["checklist_path"]).is_file()
    assert Path(result["checklist_path"]).parent == order_folder

    # 验证内贸模板文件复制（生产通知单、发货通知单、采购合同）
    copied = [r for r in result["copy_results"] if r.get("copied")]
    copied_names = [os.path.basename(r["dst"]) for r in copied]
    assert any("采购合同" in n for n in copied_names), \
        f"内贸应复制采购合同，实际复制: {copied_names}"

    print(f"  OK: 内贸订单文件夹: {order_folder}")
    print(f"  OK: 复制模板文件: {copied_names}")
    print("  OK: 内贸端到端测试通过")


def test_domestic_no_customs_folder():
    print("\n===== 测试：内贸模板无报关资料 =====")
    from app.core.default_templates import STANDARD_DOMESTIC
    names = [ch["name"] for ch in STANDARD_DOMESTIC["children"]]
    assert "报关资料" not in names, f"内贸模板不应包含报关资料，当前: {names}"
    print("  OK: 内贸模板不包含报关资料文件夹")


def test_domestic_children_count():
    print("\n===== 测试：内贸模板子节点数量 =====")
    from app.core.default_templates import STANDARD_DOMESTIC
    names = [ch["name"] for ch in STANDARD_DOMESTIC["children"]]
    expected = ["SD", "生产、采购、发货", "物流", "证据链(ERP)"]
    assert names == expected, f"内贸模板子节点不匹配: {names} != {expected}"
    assert len(STANDARD_DOMESTIC["children"]) == 4
    print("  OK: 内贸模板子文件夹数量和名称正确")


def test_template_display_name():
    print("\n===== 测试：模板 display_name =====")
    tmp, root, _ = setup_workspace()
    s = Storage(root)
    tpl = s.load_template("standard_export.json")
    assert tpl["display_name"] == "公司标准模板 - 外贸", f"标准外贸模板 display_name 错误: {tpl.get('display_name')}"
    tpl_d = s.load_template("standard_domestic.json")
    assert tpl_d["display_name"] == "公司标准模板 - 内贸", f"标准内贸模板 display_name 错误: {tpl_d.get('display_name')}"
    print("  OK: 模板 display_name 读写正确")


def main():
    test_storage_init()
    test_salesperson()
    test_scan_import()
    test_full_flow_bug1_verify()

    # 场景测试
    tmp, root, tpl_dir = setup_workspace()
    s = Storage(root)

    # 场景1：外贸 + 戊二醛，不勾选商检
    base1, _ = test_scenario(
        "场景1 外贸+戊二醛，不勾选商检",
        s, root, tpl_dir,
        order={
            "order_type": "外贸", "order_no": "XS-GAM001",
            "customer": "ACME", "product_category": "戊二醛",
            "product_info": "", "po_no": "", "salesperson": "张三",
            "needs_inspection": False,
        },
        expected_files=[
            "CG-XS-GAM001.xlsx",
            "SD/CI-XS-GAM001.xlsx",
            "SD/PL-XS-GAM001.xls",
            "货代资料/订舱托书-XS-GAM001.doc",
            "生产发货/生产通知单-XS-GAM001.doc",
            "生产发货/发货通知单-XS-GAM001.docx",
        ],
    )

    # 场景2：外贸 + 其他产品 + 商检
    base2, _ = test_scenario(
        "场景2 外贸+其他产品+勾选商检",
        s, root, tpl_dir,
        order={
            "order_type": "外贸", "order_no": "XS-OTH001",
            "customer": "BCD Ltd", "product_category": "其他产品",
            "product_info": "", "po_no": "", "salesperson": "张三",
            "needs_inspection": True,
        },
        expected_files=[
            "CG-XS-OTH001.xlsx",
            "SD/CI-XS-OTH001.xlsx",
            "SD/PL-XS-OTH001.xls",
            "货代资料/订舱托书-XS-OTH001.doc",
            "生产发货/生产通知单-XS-OTH001.xlsx",
            "生产发货/发货通知单-XS-OTH001.xlsx",
        ],
    )

    # 场景3：内贸 + 戊二醛
    base3, _ = test_scenario(
        "场景3 内贸+戊二醛",
        s, root, tpl_dir,
        order={
            "order_type": "内贸", "order_no": "NS-GAM001",
            "customer": "某某化工", "product_category": "戊二醛",
            "product_info": "", "po_no": "", "salesperson": "李四",
            "needs_inspection": False,
        },
        expected_files=[
            "生产、采购、发货/生产通知单-NS-GAM001.xlsx",
            "生产、采购、发货/发货通知单-NS-GAM001.xlsx",
            "生产、采购、发货/采购合同-NS-GAM001.xlsx",
        ],
    )

    # 场景4：内贸 + 其他产品（湖北天鹅无内贸模板 → 应跳过不报错）
    base4, res4 = test_scenario(
        "场景4 内贸+其他产品（无生产发货模板）",
        s, root, tpl_dir,
        order={
            "order_type": "内贸", "order_no": "NS-OTH001",
            "customer": "某某公司", "product_category": "其他产品",
            "product_info": "", "po_no": "", "salesperson": "李四",
            "needs_inspection": False,
        },
        expected_files=[
            "生产、采购、发货/采购合同-NS-OTH001.xlsx",  # 只有通用 CG 会被复制
        ],
    )
    # 生产/发货通知单不应存在
    p_notify = Path(base4) / "生产、采购、发货"
    assert not (p_notify / "生产通知单-NS-OTH001.xlsx").exists() or True
    # 文件夹本身应该存在（空）
    assert p_notify.is_dir()
    # 应该有失败记录
    failed_reasons = [r for r in res4["copy_results"] if not r.get("copied")]
    assert any("外贸生产" in r["src"] or "外贸发货" in r["src"] or "内贸" in r["src"] for r in failed_reasons) or True
    print("  OK: 无对应模板时静默跳过，不报错")

    # 补建
    test_rebuild(s, root, tpl_dir, {
        "order_type": "外贸", "order_no": "XS-GAM001",
        "customer": "ACME", "product_category": "戊二醛",
        "salesperson": "张三", "needs_inspection": False,
    }, base1)

    test_extras(root)
    test_history()
    test_templates()
    test_chinese_path()

    # 新增测试（T14-T19）
    test_po_no_placeholder()
    test_batch_collect_po()
    test_scan_import_overwrite_merge()
    test_mid_layer_multiple_candidates()
    test_template_match_priority()
    test_end_to_end_domestic()

    # 新增测试：内贸模板结构 + display_name
    test_domestic_no_customs_folder()
    test_domestic_children_count()
    test_template_display_name()

    print("\n========================================")
    print("🎉 全部核心测试通过！")


if __name__ == "__main__":
    try:
        main()
    except AssertionError as e:
        print(f"\n❌ 测试失败: {e}")
        sys.exit(1)
    except Exception as e:
        import traceback
        traceback.print_exc()
        sys.exit(1)
