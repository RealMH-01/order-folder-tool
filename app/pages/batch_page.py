# -*- coding: utf-8 -*-
"""批量导入页"""

import os
from datetime import datetime

from PyQt5.QtCore import Qt, pyqtSignal
from PyQt5.QtGui import QBrush, QColor
from PyQt5.QtWidgets import (
    QCheckBox, QComboBox, QFileDialog, QGridLayout, QGroupBox, QHBoxLayout,
    QHeaderView, QLabel, QLineEdit, QMessageBox, QPushButton, QSpinBox,
    QTableWidget, QTableWidgetItem, QVBoxLayout, QWidget
)

from ..core import folder_builder


HEADERS = ["订单类型", "订单号", "客户名称", "产品信息", "客户PO号", "产品类别", "是否商检", "状态"]


class BatchPage(QWidget):
    request_back = pyqtSignal()

    def __init__(self, storage, parent=None):
        super().__init__(parent)
        self.storage = storage
        self._build_ui()

    # ============== UI ==============
    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(20, 14, 20, 14)
        root.setSpacing(10)

        top = QHBoxLayout()
        btn_back = QPushButton("← 返回首页")
        btn_back.setObjectName("SecondaryButton")
        btn_back.clicked.connect(self.request_back.emit)
        top.addWidget(btn_back)
        title = QLabel("批量导入")
        title.setObjectName("TitleLabel")
        top.addWidget(title)
        top.addStretch(1)
        root.addLayout(top)

        # ----- 身份与模板 -----
        id_group = QGroupBox("① 身份与模板（批量订单共用）")
        id_layout = QGridLayout(id_group)
        id_layout.addWidget(QLabel("业务员："), 0, 0)
        self.cmb_sales = QComboBox()
        id_layout.addWidget(self.cmb_sales, 0, 1)
        id_layout.addWidget(QLabel("客户（可留空，使用每行的客户名称）："), 0, 2)
        self.cmb_customer = QComboBox()
        self.cmb_customer.setEditable(True)
        id_layout.addWidget(self.cmb_customer, 0, 3)
        tip = QLabel("提示：批量导入时，模板按「订单类型 + 业务员 + 客户」自动匹配。")
        tip.setStyleSheet("color:#666;")
        id_layout.addWidget(tip, 1, 0, 1, 4)
        self.cmb_sales.currentIndexChanged.connect(self._reload_customers)
        root.addWidget(id_group)

        # ----- 导入方式 -----
        op_group = QGroupBox("② 导入订单")
        op_layout = QHBoxLayout(op_group)
        btn_dl = QPushButton("下载 Excel 模板")
        btn_dl.setObjectName("SecondaryButton")
        btn_dl.clicked.connect(self._download_template)
        btn_import = QPushButton("导入 Excel")
        btn_import.clicked.connect(self._import_excel)
        btn_add = QPushButton("＋ 添加一行")
        btn_add.setObjectName("SecondaryButton")
        btn_add.clicked.connect(lambda: self._add_row())
        btn_del = QPushButton("－ 删除选中行")
        btn_del.setObjectName("SecondaryButton")
        btn_del.clicked.connect(self._del_rows)
        self.spin_rows = QSpinBox()
        self.spin_rows.setRange(1, 200)
        self.spin_rows.setValue(5)
        btn_gen = QPushButton("按预设行数生成")
        btn_gen.setObjectName("SecondaryButton")
        btn_gen.clicked.connect(self._gen_rows)

        op_layout.addWidget(btn_dl)
        op_layout.addWidget(btn_import)
        op_layout.addSpacing(20)
        op_layout.addWidget(btn_add)
        op_layout.addWidget(btn_del)
        op_layout.addSpacing(20)
        op_layout.addWidget(QLabel("预设行数："))
        op_layout.addWidget(self.spin_rows)
        op_layout.addWidget(btn_gen)
        op_layout.addStretch(1)
        root.addWidget(op_group)

        # ----- 表格 -----
        self.table = QTableWidget(0, len(HEADERS) + 1)  # +1 for 序号
        all_headers = ["序号"] + HEADERS
        self.table.setHorizontalHeaderLabels(all_headers)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        # 列宽：序号 50, 订单类型 80, 订单号 180, 客户名称 180, 产品信息 160,
        #      客户PO号 140, 产品类别 100, 是否商检 80, 状态 220
        self.table.setColumnWidth(0, 50)
        self.table.setColumnWidth(1, 80)
        self.table.setColumnWidth(2, 180)
        self.table.setColumnWidth(3, 180)
        self.table.setColumnWidth(4, 160)
        self.table.setColumnWidth(5, 140)
        self.table.setColumnWidth(6, 100)
        self.table.setColumnWidth(7, 80)
        self.table.setColumnWidth(8, 220)
        root.addWidget(self.table, 1)

        # ----- 执行 -----
        bottom = QHBoxLayout()
        bottom.addStretch(1)
        btn_preview = QPushButton("预览全部")
        btn_preview.setObjectName("SecondaryButton")
        btn_preview.clicked.connect(self._preview_all)
        btn_run = QPushButton("确认批量创建")
        btn_run.setStyleSheet("font-size:14px; font-weight:bold; padding:8px 24px;")
        btn_run.clicked.connect(self._run_all)
        bottom.addWidget(btn_preview)
        bottom.addWidget(btn_run)
        root.addLayout(bottom)

    # ============== 外部入口 ==============
    def refresh(self):
        self._load_salespersons()

    # ============== 数据 ==============
    def _load_salespersons(self):
        self.cmb_sales.blockSignals(True)
        self.cmb_sales.clear()
        names = [it["name"] for it in self.storage.load_salespersons()]
        self.cmb_sales.addItems([""] + names)
        self.cmb_sales.blockSignals(False)
        self._reload_customers()

    def _reload_customers(self):
        sales = self.cmb_sales.currentText()
        self.cmb_customer.clear()
        if sales:
            self.cmb_customer.addItems([""] + self.storage.get_customers(sales))
        else:
            self.cmb_customer.addItems([""])

    # ============== 表格操作 ==============
    def _add_row(self, data=None):
        """data: dict 可预填"""
        r = self.table.rowCount()
        self.table.insertRow(r)
        # 序号
        it = QTableWidgetItem(str(r + 1))
        it.setFlags(it.flags() & ~Qt.ItemIsEditable)
        it.setTextAlignment(Qt.AlignCenter)
        self.table.setItem(r, 0, it)

        # 订单类型 ComboBox
        cmb_type = QComboBox()
        cmb_type.addItems(["外贸", "内贸"])
        if data and data.get("order_type") in ("外贸", "内贸"):
            cmb_type.setCurrentText(data["order_type"])
        self.table.setCellWidget(r, 1, cmb_type)

        # 订单号（per-row 业务员存入该单元格的 UserRole，供 _collect_rows 使用）
        no_item = QTableWidgetItem(data.get("order_no", "") if data else "")
        per_row_sp = (data.get("salesperson") or "").strip() if data else ""
        no_item.setData(Qt.UserRole, per_row_sp)
        self.table.setItem(r, 2, no_item)
        # 客户
        self.table.setItem(r, 3, QTableWidgetItem(data.get("customer", "") if data else ""))
        # 产品信息
        self.table.setItem(r, 4, QTableWidgetItem(data.get("product_info", "") if data else ""))

        # 客户PO号
        self.table.setItem(r, 5, QTableWidgetItem(data.get("po_no", "") if data else ""))

        # 产品类别
        cmb_cat = QComboBox()
        cmb_cat.addItems(["戊二醛", "其他产品"])
        if data and data.get("product_category") in ("戊二醛", "其他产品"):
            cmb_cat.setCurrentText(data["product_category"])
        self.table.setCellWidget(r, 6, cmb_cat)

        # 是否商检
        chk = QCheckBox()
        w = QWidget()
        lay = QHBoxLayout(w)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setAlignment(Qt.AlignCenter)
        lay.addWidget(chk)
        if data and data.get("needs_inspection"):
            chk.setChecked(True)
        self.table.setCellWidget(r, 7, w)

        # 状态
        self.table.setItem(r, 8, QTableWidgetItem(""))

        self._renumber()

    def _del_rows(self):
        rows = sorted({i.row() for i in self.table.selectedIndexes()}, reverse=True)
        if not rows:
            QMessageBox.information(self, "提示", "请选中要删除的行")
            return
        for r in rows:
            self.table.removeRow(r)
        self._renumber()

    def _gen_rows(self):
        self.table.setRowCount(0)
        for _ in range(self.spin_rows.value()):
            self._add_row()

    def _renumber(self):
        for r in range(self.table.rowCount()):
            it = self.table.item(r, 0)
            if it:
                it.setText(str(r + 1))

    # ============== Excel ==============
    def _download_template(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "保存 Excel 模板", "批量导入模板.xlsx", "Excel 文件 (*.xlsx)")
        if not path:
            return
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = "批量导入"
        headers = ["订单类型", "订单号", "客户名称", "产品信息", "客户PO号",
                   "产品类别", "是否需要商检", "业务员"]
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2196F3")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        # 示例（业务员列为空时，使用页面顶部选择的业务员）
        ws.append(["外贸", "XS-GAM2508056NH", "ACME CO., LTD.", "戊二醛 200KG",
                   "PO-2026-001", "戊二醛", "是", "张三"])
        ws.append(["内贸", "NS-GAM2508057", "某某化工", "戊二醛 1T",
                   "", "戊二醛", "否", ""])
        widths = [12, 22, 24, 28, 18, 12, 14, 12]
        from openpyxl.utils import get_column_letter
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        wb.save(path)
        QMessageBox.information(self, "成功", f"模板已保存到：\n{path}")

    def _import_excel(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "选择 Excel 文件", "", "Excel 文件 (*.xlsx *.xls)")
        if not path:
            return
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                QMessageBox.warning(self, "提示", "Excel 为空")
                return
            header = [str(x).strip() if x else "" for x in rows[0]]
            def idx(name):
                for i, h in enumerate(header):
                    if name in h:
                        return i
                return -1
            i_type = idx("订单类型")
            i_no = idx("订单号")
            i_cust = idx("客户名称")
            i_prod = idx("产品信息")
            i_po = idx("PO号")
            i_cat = idx("产品类别")
            i_insp = idx("商检")
            i_sp = idx("业务员")

            count = 0
            for r in rows[1:]:
                if not r or all(x is None or str(x).strip() == "" for x in r):
                    continue
                data = {
                    "order_type": (str(r[i_type]).strip() if i_type >= 0 and r[i_type] else "外贸"),
                    "order_no": (str(r[i_no]).strip() if i_no >= 0 and r[i_no] else ""),
                    "customer": (str(r[i_cust]).strip() if i_cust >= 0 and r[i_cust] else ""),
                    "product_info": (str(r[i_prod]).strip() if i_prod >= 0 and r[i_prod] else ""),
                    "po_no": (str(r[i_po]).strip() if i_po >= 0 and r[i_po] else ""),
                    "product_category": (str(r[i_cat]).strip() if i_cat >= 0 and r[i_cat] else "戊二醛"),
                    "salesperson": (str(r[i_sp]).strip() if i_sp >= 0 and r[i_sp] else ""),
                }
                if data["product_category"] not in ("戊二醛", "其他产品"):
                    data["product_category"] = "戊二醛"
                if data["order_type"] not in ("外贸", "内贸"):
                    data["order_type"] = "外贸"
                insp_raw = str(r[i_insp]).strip() if i_insp >= 0 and r[i_insp] is not None else ""
                data["needs_inspection"] = insp_raw in ("是", "Y", "y", "YES", "yes", "1", "true", "True", "✓")
                self._add_row(data)
                count += 1
            QMessageBox.information(self, "成功", f"已导入 {count} 行。")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"解析 Excel 失败：{e}")

    # ============== 采集 ==============
    def _collect_rows(self):
        rows = []
        common_sales = self.cmb_sales.currentText()
        for r in range(self.table.rowCount()):
            order_type = self.table.cellWidget(r, 1).currentText()
            no_item = self.table.item(r, 2)
            order_no = no_item.text().strip() if no_item else ""
            per_row_sp = (no_item.data(Qt.UserRole) if no_item else "") or ""
            customer = self.table.item(r, 3).text().strip() if self.table.item(r, 3) else ""
            product_info = self.table.item(r, 4).text().strip() if self.table.item(r, 4) else ""
            po_no = self.table.item(r, 5).text().strip() if self.table.item(r, 5) else ""
            product_category = self.table.cellWidget(r, 6).currentText()
            chk_w = self.table.cellWidget(r, 7)
            chk = chk_w.findChild(QCheckBox) if chk_w else None
            needs_inspection = bool(chk and chk.isChecked())
            if not order_no:
                continue
            rows.append({
                "row_index": r,
                "order_type": order_type,
                "order_no": order_no,
                "customer": customer or self.cmb_customer.currentText(),
                "product_info": product_info,
                "po_no": po_no,
                "product_category": product_category,
                "needs_inspection": needs_inspection and order_type == "外贸",
                # per-row 业务员优先（Excel 中填了业务员列），否则用页面顶部的共用业务员
                "salesperson": (per_row_sp or common_sales).strip(),
            })
        return rows

    def _set_status(self, row, text, color="#333"):
        it = QTableWidgetItem(text)
        it.setForeground(QBrush(QColor(color)))
        self.table.setItem(row, 8, it)

    # ============== 预览 & 执行 ==============
    def _preview_all(self):
        rows = self._collect_rows()
        if not rows:
            QMessageBox.information(self, "提示", "没有有效订单行")
            return
        for od in rows:
            try:
                base_path = self.storage.build_customer_dir(
                    od["salesperson"], od["customer"])
                if not base_path:
                    self._set_status(od["row_index"], "路径无效", "#E53935")
                    continue
                _, tpl = self.storage.match_template(od["salesperson"], od["customer"], od["order_type"])
                if not tpl:
                    self._set_status(od["row_index"], "无可用模板", "#E53935")
                    continue
                ctx = folder_builder.build_context(od)
                tpl_folders = folder_builder.flatten_template_folders(
                    tpl, base_path, ctx, od["needs_inspection"])
                tpl_folders, _ = folder_builder.compare_with_existing(
                    base_path, tpl_folders)
                exists = sum(1 for i in tpl_folders if i["status"] == "existing" and not i.get("is_root"))
                tocre = sum(1 for i in tpl_folders if i["status"] == "to_create" and not i.get("is_root"))
                root_existed = any(i["status"] == "existing" and i.get("is_root") for i in tpl_folders)
                if root_existed:
                    self._set_status(od["row_index"], f"待补建 {tocre}，已存在 {exists}", "#1976D2")
                else:
                    self._set_status(od["row_index"], f"待创建 {tocre} 个", "#1976D2")
            except Exception as e:
                self._set_status(od["row_index"], f"错误：{e}", "#E53935")

    def _run_all(self):
        rows = self._collect_rows()
        if not rows:
            QMessageBox.information(self, "提示", "没有有效订单行")
            return
        reply = QMessageBox.question(self, "确认",
                                     f"即将批量创建 {len(rows)} 笔订单文件夹，是否继续？")
        if reply != QMessageBox.Yes:
            return
        cfg = self.storage.load_config()
        tpl_dir = cfg.get("template_files_dir") or None
        success, fail = 0, 0
        details = []
        for od in rows:
            try:
                base_path = self.storage.build_customer_dir(
                    od["salesperson"], od["customer"])
                if not base_path:
                    self._set_status(od["row_index"], "路径无效", "#E53935")
                    fail += 1
                    continue
                tpl_name, tpl = self.storage.match_template(od["salesperson"], od["customer"], od["order_type"])
                if not tpl:
                    self._set_status(od["row_index"], "无可用模板", "#E53935")
                    fail += 1
                    continue
                result = folder_builder.execute_build(
                    order=od, template=tpl, base_path=base_path,
                    template_files_dir=tpl_dir)
                # 历史
                self.storage.append_history({
                    "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "operator": cfg.get("operator", ""),
                    "salesperson": od["salesperson"],
                    "customer": od["customer"],
                    "order_no": od["order_no"],
                    "order_type": od["order_type"],
                    "product_category": od["product_category"],
                    "template_name": tpl_name,
                    "path": result["base_path"],
                    "result": "成功",
                    "created_count": len(result["created"]),
                    "skipped_count": len(result["skipped"]),
                    "copied_count": sum(1 for r in result["copy_results"] if r.get("copied")),
                })
                self._set_status(od["row_index"],
                                 f"✅ 新建 {len(result['created'])}，跳过 {len(result['skipped'])}，复制模板 {sum(1 for r in result['copy_results'] if r.get('copied'))}",
                                 "#2E7D32")
                success += 1
                details.append(f"[{od['order_no']}] 新建 {len(result['created'])} / 跳过 {len(result['skipped'])} / 复制 {sum(1 for r in result['copy_results'] if r.get('copied'))}")
            except Exception as e:
                self._set_status(od["row_index"], f"❌ {e}", "#E53935")
                fail += 1
                details.append(f"[{od['order_no']}] 失败：{e}")

        # 汇总
        from PyQt5.QtWidgets import QDialog, QPlainTextEdit, QPushButton, QVBoxLayout
        dlg = QDialog(self)
        dlg.setWindowTitle("批量执行完成")
        dlg.resize(700, 460)
        v = QVBoxLayout(dlg)
        v.addWidget(QLabel(f"<b>成功 {success} 笔，失败 {fail} 笔。</b>"))
        tx = QPlainTextEdit()
        tx.setReadOnly(True)
        tx.setPlainText("\n".join(details))
        v.addWidget(tx)
        btn = QPushButton("关闭")
        btn.clicked.connect(dlg.accept)
        v.addWidget(btn)
        dlg.exec_()
