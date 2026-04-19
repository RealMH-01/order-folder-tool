# -*- coding: utf-8 -*-
"""
订单文件夹生成器：
- 扫描目标路径已存在的子文件夹，与模板对比
- 创建缺失的文件夹
- 按模板复制并重命名模板文件
- 在订单根目录生成 Excel 勾选表
"""

import os
import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ------------------------------------------------------------------
# 占位符替换
# ------------------------------------------------------------------
PLACEHOLDER_RE = re.compile(r"<[^<>]+>")


def replace_placeholders(text: str, ctx: Dict[str, str]) -> str:
    """把 <订单号>/<客户名称>/<日期>/<业务员>/<客户PO号>/<产品信息> 等替换"""
    if not text:
        return text

    def _sub(m):
        key = m.group(0)
        return ctx.get(key, key)
    return PLACEHOLDER_RE.sub(_sub, text)


def build_context(order: Dict[str, Any]) -> Dict[str, str]:
    """根据订单信息构造占位符上下文"""
    return {
        "<订单号>": order.get("order_no", ""),
        "<客户名称>": order.get("customer", ""),
        "<客户PO号>": order.get("po_no", ""),
        "<产品信息>": order.get("product_info", ""),
        "<日期>": datetime.now().strftime("%Y%m%d"),
        "<业务员>": order.get("salesperson", ""),
        "<SHXY编号>": order.get("shxy_no", "SHXY"),
    }


# ------------------------------------------------------------------
# 产地匹配
# ------------------------------------------------------------------
# product_category: "戊二醛" → 宁夏；"其他产品" → 湖北天鹅
ORIGIN_MAP = {
    "戊二醛": "宁夏",
    "其他产品": "湖北天鹅",
}

# 宁夏、湖北天鹅的文件名扩展名约定（参考需求 §9.1）
ORIGIN_FILE_EXT = {
    ("宁夏", "外贸生产"): ".doc",
    ("宁夏", "外贸发货"): ".docx",
    ("宁夏", "内贸生产"): ".xlsx",
    ("宁夏", "内贸发货"): ".xlsx",
    ("湖北天鹅", "外贸生产"): ".xlsx",
    ("湖北天鹅", "外贸发货"): ".xlsx",
    ("湖北天鹅", "内贸生产"): ".xlsx",  # 需求中"其他产品+内贸"没有模板，但为安全起见保留
    ("湖北天鹅", "内贸发货"): ".xlsx",
}


def resolve_file_template(tmpl: Optional[str], product_category: str) -> Optional[str]:
    """把 file_template 中的 [产地] 标记替换成实际子目录+文件名。

    若无 [产地] 标记，直接返回 tmpl。
    """
    if not tmpl:
        return tmpl
    if "[产地]" not in tmpl:
        return tmpl

    origin = ORIGIN_MAP.get(product_category)
    if not origin:
        return None  # 未知产品类别

    # 形如 "[产地]外贸生产" → "宁夏/宁夏外贸生产.doc"
    suffix = tmpl.replace("[产地]", "")  # 如 "外贸生产"
    ext = ORIGIN_FILE_EXT.get((origin, suffix))
    if not ext:
        return None
    return f"{origin}/{origin}{suffix}{ext}"


def resolve_filename_with_ext(filename: str, file_template: Optional[str],
                              product_category: str) -> str:
    """
    ref_files 里 filename 可能没有后缀（如"生产通知单-<订单号>"），
    这时需要从模板文件路径取扩展名。
    """
    if "." in os.path.basename(filename):
        return filename
    # 没有扩展名 → 看模板
    resolved = resolve_file_template(file_template, product_category)
    if resolved:
        ext = os.path.splitext(resolved)[1]
        if ext:
            return filename + ext
    return filename


# ------------------------------------------------------------------
# 扫描比对
# ------------------------------------------------------------------
def flatten_template_folders(template: Dict[str, Any],
                             base_path: str,
                             ctx: Dict[str, str],
                             needs_inspection: bool,
                             parent_path: Optional[str] = None) -> List[Dict[str, Any]]:
    """
    把模板 JSON 展开为"要创建的文件夹列表"。

    - 根节点（parent_path is None）：rel_path=<订单号名称>，is_root=True
      对应客户目录下的"订单号总文件夹"，所有子节点都放在它里面。
    - 子节点：rel_path 是相对 base_path 的路径，如 "<订单号>/SD"、"<订单号>/货代资料/唛头"

    base_path 是客户目录（不含订单号）。
    """
    result = []
    name = replace_placeholders(template.get("name", ""), ctx)

    # 判断 optional / condition
    if template.get("optional"):
        condition = template.get("condition")
        if condition == "needs_inspection" and not needs_inspection:
            return result

    # 未被勾选的节点直接过滤（模板编辑器会给 _enabled=False）
    if template.get("_enabled") is False:
        return result

    is_root = (parent_path is None)
    # 修复：根节点也要有 rel_path（即订单号文件夹名），
    # 这样所有子文件夹才会真正创建在"订单号文件夹"内部，
    # 而不是散落在客户目录下。
    if is_root:
        rel_path = name
    else:
        rel_path = os.path.join(parent_path, name) if parent_path else name

    result.append({
        "rel_path": rel_path,
        "name": name,
        "ref_files": template.get("ref_files", []) or [],
        "is_root": is_root,
    })

    for child in template.get("children", []) or []:
        result.extend(flatten_template_folders(
            child, base_path, ctx, needs_inspection,
            parent_path=rel_path,
        ))
    return result


def compare_with_existing(base_path: str,
                          template_folders: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    扫描 base_path 是否已存在，对每个待创建节点与磁盘对比。
    base_path 为客户目录（不含订单号），订单号文件夹是模板根节点。
    返回带 status 字段的列表：existing | to_create | out_of_template
    同时扫描订单号文件夹下磁盘上但不在模板里的目录。
    """
    base = Path(base_path)
    result = []
    template_paths = set()
    # 订单号根文件夹相对路径（用于扫描 extras 时的 scan 根）
    order_root_rel = ""
    for item in template_folders:
        if item.get("is_root"):
            order_root_rel = item["rel_path"]
            break

    for item in template_folders:
        rel = item["rel_path"]
        status = "to_create"
        if base.exists():
            abs_path = base / rel if rel else base
            if abs_path.exists() and abs_path.is_dir():
                status = "existing"
            else:
                status = "to_create"
        template_paths.add(rel)
        new_item = dict(item)
        new_item["status"] = status
        result.append(new_item)

    # 扫描磁盘上额外目录（仅扫描订单号根目录下的子目录，避免把客户目录下的
    # 其他订单文件夹误认为"模板外"）
    extras = []
    order_root_abs = base / order_root_rel if order_root_rel else base
    if order_root_abs.exists():
        extras = scan_extra_folders(base, order_root_abs, template_paths)
    return result, extras


def scan_extra_folders(root: Path, base: Path,
                        template_paths: set) -> List[Dict[str, Any]]:
    """递归扫描 base 下所有目录，返回不在 template_paths 中的"""
    extras = []
    try:
        for p in base.iterdir():
            if p.is_dir() and not p.name.startswith("."):
                rel = str(p.relative_to(root)).replace("\\", "/")
                # 注意 Windows 上分隔符
                rel_os = os.path.relpath(str(p), str(root))
                if rel_os not in template_paths:
                    extras.append({
                        "rel_path": rel_os,
                        "name": p.name,
                        "status": "out_of_template",
                    })
                extras.extend(scan_extra_folders(root, p, template_paths))
    except PermissionError:
        pass
    return extras


# ------------------------------------------------------------------
# 创建 & 复制
# ------------------------------------------------------------------
def create_folders(base_path: str,
                   template_folders: List[Dict[str, Any]]) -> Tuple[List[str], List[str]]:
    """创建所有 to_create 节点，返回 (created, skipped)。

    base_path 是客户目录（不含订单号）。订单号文件夹作为根节点也会被创建。
    """
    created, skipped = [], []
    base = Path(base_path)
    base.mkdir(parents=True, exist_ok=True)
    for item in template_folders:
        rel = item["rel_path"]
        target = base / rel if rel else base
        if target.exists():
            skipped.append(str(target))
        else:
            target.mkdir(parents=True, exist_ok=True)
            created.append(str(target))
    return created, skipped


def copy_template_files(base_path: str,
                        template_folders: List[Dict[str, Any]],
                        template_files_dir: Optional[str],
                        ctx: Dict[str, str],
                        product_category: str) -> List[Dict[str, Any]]:
    """
    复制模板文件到目标文件夹并重命名。
    返回复制结果列表：
    [{"folder": "...", "src": "...", "dst": "...", "copied": True/False, "reason": "..."}, ...]
    """
    results = []
    if not template_files_dir or not os.path.isdir(template_files_dir):
        return results

    base = Path(base_path)
    tpl_dir = Path(template_files_dir)

    for item in template_folders:
        rel = item["rel_path"]
        target_folder = base / rel if rel else base
        for rf in item.get("ref_files", []):
            file_template = rf.get("file_template")
            if not file_template:
                continue
            resolved = resolve_file_template(file_template, product_category)
            if not resolved:
                results.append({
                    "folder": str(target_folder),
                    "src": file_template,
                    "dst": "",
                    "copied": False,
                    "reason": "未找到匹配的产地模板",
                })
                continue
            src = tpl_dir / resolved
            if not src.exists():
                results.append({
                    "folder": str(target_folder),
                    "src": str(src),
                    "dst": "",
                    "copied": False,
                    "reason": "模板文件不存在",
                })
                continue
            # 目标文件名 = ref_files.filename 按占位符替换；若没后缀则用模板文件扩展名
            raw_name = rf.get("filename", "")
            filled_name = replace_placeholders(raw_name, ctx)
            if "." not in os.path.basename(filled_name):
                filled_name += src.suffix
            # 功能 C：复制的空白模板文件自动添加 "_对照" 后缀，
            # 用于与用户实际工作文件区分，便于后续「整理文件夹」识别删除。
            name_part, ext_part = os.path.splitext(filled_name)
            filled_name = f"{name_part}_对照{ext_part}"
            dst = target_folder / filled_name
            try:
                target_folder.mkdir(parents=True, exist_ok=True)
                shutil.copy2(str(src), str(dst))
                results.append({
                    "folder": str(target_folder),
                    "src": str(src),
                    "dst": str(dst),
                    "copied": True,
                    "reason": "",
                })
            except Exception as e:
                results.append({
                    "folder": str(target_folder),
                    "src": str(src),
                    "dst": str(dst),
                    "copied": False,
                    "reason": f"复制失败: {e}",
                })
    return results


# ------------------------------------------------------------------
# Excel 勾选表
# ------------------------------------------------------------------
def generate_checklist_excel(order_folder_path: str,
                              order_no: str,
                              template_folders: List[Dict[str, Any]],
                              copy_results: List[Dict[str, Any]],
                              ctx: Dict[str, str],
                              product_category: str,
                              base_path: Optional[str] = None) -> str:
    """在订单根目录（order_folder_path）生成"文件清单-<订单号>.xlsx"，返回文件路径。

    template_folders 中的 rel_path 是相对 base_path（客户目录）的路径；
    如果未传入 base_path，则默认等于 order_folder_path（向后兼容）。
    """
    if base_path is None:
        base_path = order_folder_path
    wb = Workbook()
    ws = wb.active
    ws.title = "文件清单"

    # 样式
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2196F3")
    group_font = Font(name="微软雅黑", size=11, bold=True)
    group_fill = PatternFill("solid", fgColor="BBDEFB")
    normal_font = Font(name="微软雅黑", size=10)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="B0BEC5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 表头
    headers = ["✓", "文件夹", "文件名", "来源", "有模板", "备注"]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws.freeze_panes = "A2"

    # 构造"文件夹 → 是否已复制某文件" 的索引
    copied_set = set()
    for r in copy_results:
        if r.get("copied"):
            copied_set.add(r["dst"])

    row = 2
    for item in template_folders:
        ref_files = item.get("ref_files", [])
        if not ref_files:
            continue
        # 小标题行
        folder_display = item["name"] if not item["is_root"] else f"根目录（{order_no}）"
        ws.cell(row=row, column=1, value="").border = border
        ws.cell(row=row, column=2, value=folder_display).font = group_font
        ws.cell(row=row, column=2).fill = group_fill
        ws.cell(row=row, column=2).alignment = left
        # 合并 B..F 作为标题
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        for c in range(1, 7):
            ws.cell(row=row, column=c).border = border
        row += 1

        rel = item["rel_path"]
        target_folder = Path(base_path) / rel if rel else Path(base_path)

        for rf in ref_files:
            raw_name = rf.get("filename", "")
            file_template = rf.get("file_template")
            filled_name = replace_placeholders(raw_name, ctx)
            resolved_tpl = resolve_file_template(file_template, product_category)
            # 对无后缀文件名补上模板文件的扩展名
            if "." not in os.path.basename(filled_name) and resolved_tpl:
                filled_name += os.path.splitext(resolved_tpl)[1]
            has_template = "是" if file_template else "否"
            source = rf.get("source", "")

            # 功能 C：如果该文件是从模板复制出来的空白对照文件，
            # 文件名需带 "_对照" 后缀，与 copy_template_files 保持一致。
            display_name = filled_name
            if file_template and resolved_tpl:
                name_part, ext_part = os.path.splitext(filled_name)
                display_name = f"{name_part}_对照{ext_part}"

            # 是否已复制
            dst_path = str(target_folder / display_name)
            already_copied = dst_path in copied_set
            check_mark = "✓" if already_copied else ""
            remark = "已从模板复制" if already_copied else ""

            ws.cell(row=row, column=1, value=check_mark).alignment = center
            ws.cell(row=row, column=2, value=folder_display).alignment = left
            ws.cell(row=row, column=3, value=display_name).alignment = left
            ws.cell(row=row, column=4, value=source).alignment = left
            ws.cell(row=row, column=5, value=has_template).alignment = center
            ws.cell(row=row, column=6, value=remark).alignment = left
            for c in range(1, 7):
                cell = ws.cell(row=row, column=c)
                cell.border = border
                if not cell.font or cell.font != group_font:
                    cell.font = normal_font
            row += 1

    # 列宽自适应
    widths = [6, 24, 36, 22, 10, 24]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    out_path = Path(order_folder_path) / f"文件清单-{order_no}.xlsx"
    wb.save(str(out_path))
    return str(out_path)


# ------------------------------------------------------------------
# 顶层：执行整套创建
# ------------------------------------------------------------------
def execute_build(order: Dict[str, Any],
                  template: Dict[str, Any],
                  base_path: str,
                  template_files_dir: Optional[str]) -> Dict[str, Any]:
    """
    执行整套流程：
    1. 展开模板
    2. 创建缺失文件夹
    3. 复制模板文件
    4. 生成 Excel 清单

    返回详细结果字典。
    """
    ctx = build_context(order)
    needs_inspection = bool(order.get("needs_inspection", False))
    product_category = order.get("product_category", "")
    order_no = order.get("order_no", "")

    template_folders = flatten_template_folders(
        template, base_path, ctx, needs_inspection
    )

    created, skipped = create_folders(base_path, template_folders)

    copy_results = copy_template_files(
        base_path, template_folders, template_files_dir, ctx, product_category
    )

    # 订单号根目录（客户目录下的订单号文件夹）
    order_root_rel = ""
    for it in template_folders:
        if it.get("is_root"):
            order_root_rel = it["rel_path"]
            break
    order_folder_path = str(Path(base_path) / order_root_rel) if order_root_rel else base_path

    checklist_path = ""
    try:
        checklist_path = generate_checklist_excel(
            order_folder_path, order_no, template_folders, copy_results,
            ctx, product_category, base_path=base_path
        )
    except Exception as e:
        checklist_path = f"生成Excel清单失败: {e}"

    return {
        # 订单号总文件夹路径（即"打开订单文件夹"应该打开的位置）
        "base_path": order_folder_path,
        # 客户目录路径（不含订单号）
        "customer_dir": base_path,
        "created": created,
        "skipped": skipped,
        "copy_results": copy_results,
        "checklist_path": checklist_path,
    }
